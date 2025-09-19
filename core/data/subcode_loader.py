from oauth2client.service_account import ServiceAccountCredentials
import gspread
from googleapiclient.discovery import build
import pandas as pd
from datetime import datetime, timedelta, date
from openpyxl.utils import get_column_letter
import os
import re
import io
import time
from decimal import Decimal
try:
    # 新構造のログ管理を優先使用
    from src.core.logging.logger import get_logger
    LOGGER = get_logger('datasets')
except ImportError:
    # フォールバック：旧構造
    from core.config.my_logging import setup_department_logger
    LOGGER = setup_department_logger('datasets', app_type='datasets')
from tenacity import retry, stop_after_attempt, wait_exponential, before_sleep_log
import shutil
import traceback
import pyarrow as pa
import pyarrow.parquet as pq

# LOGGER は上記のtryブロックで設定済み

# ログシステム初期化完了

# Googleの認証処理
def authenticate_google_api(json_keyfile_path, scopes):
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scopes)
    return credentials

# リトライ処理
def retry_on_exception(func):
    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=60, min=60, max=300),
        before_sleep=before_sleep_log
    )
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)
    return wrapper

def before_sleep_log(retry_state):
    LOGGER.warning(f"Retrying {retry_state.fn.__name__} after exception: {retry_state.outcome.exception()}")

@retry_on_exception
def load_sql_file_list_from_spreadsheet(spreadsheet_id, sheet_name, json_keyfile_path, execution_column):
    """
    指定されたGoogleスプレッドシートからSQLファイルのリストを読み込む関数

    Args:
        spreadsheet_id: スプレッドシートのID。
        sheet_name: 読み込むシートの名前。
        json_keyfile_path: Google APIの認証情報が含まれるJSONファイルのパス。
        execution_column: 実行対象の列名。
        
    Returns:
        実行対象とマークされたSQLファイル名のリスト。
    """
    SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    SQL_FILE_COLUMN = 'sqlファイル名'
    CSV_FILE_COLUMN = 'CSVファイル名/SSシート名'
    FILENAME_FORMAT_COLUMN = '保存ファイル名形式'
    PERIOD_CONDITION_COLUMN = '取得期間'
    PERIOD_CRITERIA_COLUMN = '取得基準'
    SPREADSHEET_COLUMN = '出力先'
    SAVE_PATH_COLUMN = '保存先PATH/ID'
    DELETION_EXCLUSION_COLUMN = '削除R除外'
    PASTE_FORMAT_COLUMN = 'スプシ貼り付け形式'
    TEST_COLUMN = 'テスト'
    CATEGORY_COLUMN = 'カテゴリ'
    MAIN_TABLE_COLUMN = 'メインテーブル'
    CSV_FILE_NAME_COLUMN = 'CSVファイル呼称'
    SHEET_NAME_COLUMN = 'シート名'
    EXECUTION_FREQUENCY_COLUMN = '実行頻度'

    credentials = authenticate_google_api(json_keyfile_path, SCOPES)
    gc = gspread.authorize(credentials)

    try:
        worksheet = gc.open_by_key(spreadsheet_id).worksheet(sheet_name)
        records = worksheet.get_all_records()
        LOGGER.info(f"Loaded sheet: {sheet_name} with {len(records)} records.")
    except Exception as e:
        LOGGER.error(f"Failed to load worksheet '{sheet_name}' from spreadsheet '{spreadsheet_id}': {e}")
        raise

    sql_and_csv_files = []
    today = datetime.now().date()

    for record in records:
        if record.get(execution_column, '').upper() == 'TRUE':
            execution_frequency = record.get(EXECUTION_FREQUENCY_COLUMN, '')
            
            # テスト用の日付指定（コメントアウトを解除して使用）
            #today = datetime.strptime('2025-07-24', '%Y-%m-%d').date()  # テスト用日付
            today = datetime.now().date()  # 本番用
            
            # 月初の場合、1日以外はスキップ
            if execution_frequency == '月初' and today.day != 1: #テストの場合は数字を今日の日にする
                LOGGER.info(f"月初実行のため、スキップします: {record.get(SQL_FILE_COLUMN)}")
                LOGGER.info(f"実行日: {today}")
                continue

            sql_file_name = record.get(SQL_FILE_COLUMN, '')
            csv_file_name = record.get(CSV_FILE_COLUMN, '')
            filename_format = record.get(FILENAME_FORMAT_COLUMN, '')
            period_condition = record.get(PERIOD_CONDITION_COLUMN, '')
            period_criteria = record.get(PERIOD_CRITERIA_COLUMN, '')
            save_path_id = record.get(SAVE_PATH_COLUMN, '')
            output_to_spreadsheet = record.get(SPREADSHEET_COLUMN, '')
            deletion_exclusion = record.get(DELETION_EXCLUSION_COLUMN, '')
            paste_format = record.get(PASTE_FORMAT_COLUMN, '')
            test_execution = record.get(TEST_COLUMN, '')
            category = record.get(CATEGORY_COLUMN, '')
            main_table_name = record.get(MAIN_TABLE_COLUMN, '')
            csv_file_name_column = record.get(CSV_FILE_NAME_COLUMN, '')
            sheet_name_record = record.get(SHEET_NAME_COLUMN, '')

            if filename_format:
                now = datetime.now()
                filename_without_ext = os.path.splitext(csv_file_name)[0]  # 拡張子を除いたファイル名
                filename_patterns = {
                    'yyyyMMdd_{filename}': now.strftime('%Y%m%d_') + filename_without_ext,
                    '{filename}_yyyy-MM-dd': filename_without_ext + '_' + now.strftime('%Y-%m-%d'),
                    '{filename}yyyy-MM-dd': filename_without_ext + now.strftime('%Y-%m-%d'),
                    '{filename}_yyyyMMddhhmmss': filename_without_ext + '_' + now.strftime('%Y%m%d%H%M%S'),
                    '{filename}_yyyy-MM': filename_without_ext + '_' + now.strftime('%Y-%m')
                }
                for pattern, value in filename_patterns.items():
                    if pattern in filename_format:
                        csv_file_name = filename_format.replace(pattern, value) + '.csv'
                        break
                else:
                    csv_file_name = filename_format + '.csv'
            
            sql_and_csv_files.append((
                sql_file_name, csv_file_name, period_condition, period_criteria, save_path_id, 
                output_to_spreadsheet, deletion_exclusion, paste_format, test_execution, 
                category, main_table_name, csv_file_name_column, sheet_name_record
            ))

    LOGGER.info(f"処理対象件数: {len(sql_and_csv_files)}件")
    return sql_and_csv_files

@retry_on_exception
def load_sql_from_file(file_path, google_folder_id, json_keyfile_path):
    try:
        LOGGER.info(f"SQLファイル読み込み開始: {file_path}")
        SCOPES = ['https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, SCOPES)
        service = build('drive', 'v3', credentials=credentials)

        LOGGER.info(f"GoogleドライブのフォルダID: {google_folder_id}")
        LOGGER.info(f"SQLファイル名: {file_path}")

        query = f"'{google_folder_id}' in parents and name = '{file_path}'"
        file_list_results = service.files().list(
            q=query, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        files = file_list_results.get('files', [])

        LOGGER.info(f"ファイルリスト結果: {files}")

        if files:
            file_id = files[0]['id']
            LOGGER.info(f"ファイルID取得成功: {file_id}")
            file_content = service.files().get_media(fileId=file_id, supportsAllDrives=True).execute().decode('utf-8')
            LOGGER.info(f"SQLファイル読み込み成功 - 文字数: {len(file_content)}")
            LOGGER.info(f"SQL内容（最初の200文字）: {file_content[:200]}...")
            sql_query = file_content.strip()
            return sql_query
        else:
            error_message = f"ファイルが見つかりません: {file_path}"
            LOGGER.error(error_message)
            return None
    except Exception as e:
        error_message = f"SQLファイルの読み込み中にエラーが発生しました: {e}"
        LOGGER.error(error_message)
        LOGGER.error(traceback.format_exc())
        return None

# CSVファイルに保存する処理
def save_chunk_to_csv(chunk, file_path, include_header=True):
    mode = 'w' if include_header else 'a'  # ヘッダを含める場合は'w'、含めない場合は'a'
    header = include_header
    try:
        with open(file_path, mode=mode, newline='', encoding='cp932', errors='replace') as file:
            chunk.to_csv(file, index=False, header=header)
    except Exception as e:
        LOGGER.error(f"ファイル書き込み時にエラーが発生しました: {e}")
        raise

# データフレームをチャンクに分割して処理する関数
def process_dataframe_in_chunks(df, chunk_size, file_path, delay=None):
    if chunk_size is None or len(df) <= chunk_size:
        try:
            save_chunk_to_csv(df, file_path)
            LOGGER.info(f"Saved {len(df)} records to {file_path}.")
            return len(df)
        except Exception as e:
            LOGGER.error(f"ファイル書き込み時にエラーが発生しました: {e}")
            raise
    else:
        chunk_dir = os.path.join(os.path.dirname(file_path), 'chunk')
        os.makedirs(chunk_dir, exist_ok=True)
        chunk_file_paths = []
        total_records = 0
        for i, chunk in enumerate(pd.read_csv(io.StringIO(df.to_csv(index=False, encoding='cp932', sep=',')), chunksize=chunk_size, dtype=str)):
            chunk_file_path = os.path.join(chunk_dir, f"{os.path.basename(file_path)[:-4]}_chunk_{i}.csv")
            try:
                save_chunk_to_csv(chunk, chunk_file_path)
                chunk_file_paths.append(chunk_file_path)
                chunk_records = len(chunk)
                total_records += chunk_records
                if delay:
                    LOGGER.info(f"Waiting for {delay} seconds before processing the next chunk.")
                    time.sleep(delay)
            except Exception as e:
                LOGGER.error(f"チャンクファイル書き込み時にエラーが発生しました: {e}")
                raise

        combine_chunk_files(chunk_file_paths, file_path)
        shutil.rmtree(chunk_dir)
        
        LOGGER.info(f"Combined {total_records} records into {file_path}.")
        return total_records

def combine_chunk_files(chunk_file_paths, output_file_path):
    try:
        with open(output_file_path, 'w', encoding='cp932', errors='replace', newline='') as f_out:
            with open(chunk_file_paths[0], 'r', encoding='cp932', errors='replace') as f_in:
                lines = f_in.readlines()
                f_out.writelines(lines)  # ヘッダ行とデータ行を書き込む
            for file_path in chunk_file_paths[1:]:
                with open(file_path, 'r', encoding='cp932', errors='replace') as f_in:
                    lines = f_in.readlines()
                    f_out.writelines(lines[1:])  # ヘッダ行をスキップしてデータ行を書き込む
        LOGGER.info(f"Combined {len(chunk_file_paths)} chunk files into {output_file_path}.")
    except Exception as e:
        LOGGER.error(f"チャンクファイルの結合時にエラーが発生しました: {e}")
        raise

# CSVファイル処理
@retry_on_exception
def csvfile_export(conn, sql_query, csv_file_path, main_table_name, category, json_keyfile_path, spreadsheet_id, csv_file_name, csv_file_name_column, sheet_name, chunk_size=None, delay=None):
    try:
        if sheet_name:
            try:
                scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scope)
                gc = gspread.authorize(credentials)
                worksheet = gc.open_by_key(spreadsheet_id).worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound as e:
                LOGGER.error(f"ワークシート '{sheet_name}' が見つかりませんでした: {e}")
                write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
                raise

        data_types = get_data_types(worksheet) if sheet_name else {}
        df = pd.read_sql(sql_query, conn)

        # NaN、None、'nan'、'None'を空文字列に置換
        df = df.fillna('').replace({'None': '', 'nan': ''})

        # Int64型のカラムで空文字列を0に置換
        int64_columns = [col for col, dtype in data_types.items() if dtype == 'int' and col in df.columns]
        for col in int64_columns:
            df[col] = df[col].replace('', 0)

        # データ型を適用
        df = apply_data_types_to_df(df, data_types, LOGGER)

        # 数値型の列で空文字列になっているセルを0に置換（Int64型以外の数値型カラム用）
        numeric_columns = df.select_dtypes(include=['float64']).columns
        for col in numeric_columns:
            df[col] = df[col].replace('', 0)

        # データフレームの各要素を文字列に変換
        df = df.map(lambda x: str(int(x)) if isinstance(x, (float, Decimal)) and x.is_integer() else str(x) if not pd.isna(x) else '')

        # 一時ファイルパスを作成
        temp_file_path = csv_file_path + '.temp'

        try:
            record_count = process_dataframe_in_chunks(df, chunk_size, temp_file_path, delay=delay)
            
            # 処理が成功したら、一時ファイルを正式なファイルに置き換え
            if os.path.exists(csv_file_path):
                os.remove(csv_file_path)
            os.rename(temp_file_path, csv_file_path)

            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "成功", None, csv_file_path)
            LOGGER.info(f"CSVファイルが正常に保存されました: {csv_file_path} ({record_count} レコード)")
        except Exception as e:
            LOGGER.error(f"CSVファイル書き込み時にエラーが発生しました: {e}")
            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
            
            # エラーが発生した場合、一時ファイルを削除
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            
            raise

    except Exception as e:
        LOGGER.error(f"クエリ実行またはCSVファイル書き込み時にエラーが発生しました: {e}")
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
        raise

@retry_on_exception
def parquetfile_export(conn, sql_query, parquet_file_path, main_table_name, category, json_keyfile_path, spreadsheet_id, parquet_file_name, csv_file_name_column, sheet_name, chunk_size=None, delay=None):
    try:
        if sheet_name:
            try:
                scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scope)
                gc = gspread.authorize(credentials)
                worksheet = gc.open_by_key(spreadsheet_id).worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound as e:
                LOGGER.error(f"ワークシート '{sheet_name}' が見つかりませんでした: {e}")
                write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), parquet_file_path)
                raise

        data_types = get_data_types(worksheet) if sheet_name else {}
        LOGGER.info(f"Detected data types: {data_types}")
        
        df = pd.read_sql(sql_query, conn)

        LOGGER.info(f"Original DataFrame loaded with {len(df)} records.")

        # NaN、None、'nan'、'None'を空文字列に置換
        df = df.fillna('').replace({'None': '', 'nan': ''})

        # データ型を適用（Parquet用に安全な変換）
        df = apply_data_types_to_df_for_parquet(df, data_types, LOGGER)

        LOGGER.info(f"DataFrame info after type conversion:\n{df.info()}")

        # ディレクトリ作成（Windows UNCパス対応）
        os.makedirs(os.path.dirname(parquet_file_path), exist_ok=True)
        
        # 一時ファイルパスを作成（Windows UNCパス対応）
        temp_file_path = os.path.join(os.path.dirname(parquet_file_path), os.path.basename(parquet_file_path) + '.temp')

        try:
            # DataFrameをParquetファイルとして保存
            table = pa.Table.from_pandas(df)
            pq.write_table(table, temp_file_path)
            
            # 処理が成功したら、一時ファイルを正式なファイルに置き換え
            if os.path.exists(parquet_file_path):
                os.remove(parquet_file_path)
            os.rename(temp_file_path, parquet_file_path)

            record_count = len(df)
            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "成功", None, parquet_file_path)
            LOGGER.info(f"Parquetファイルが正常に保存されました: {parquet_file_path} ({record_count} レコード)")
        except Exception as e:
            LOGGER.error(f"Parquetファイル書き込み時にエラーが発生しました: {e}")
            LOGGER.error(f"エラーの詳細:\n{traceback.format_exc()}")
            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), parquet_file_path)
            
            # エラーが発生した場合、一時ファイルを削除
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            
            raise

    except Exception as e:
        LOGGER.error(f"クエリ実行またはParquetファイル書き込み時にエラーが発生しました: {e}")
        LOGGER.error(f"エラーの詳細:\n{traceback.format_exc()}")
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), parquet_file_path)
        raise

# 複数のパターンにマッチする正規表現を定義
def extract_columns_mapping(sql_query):
    patterns = [
        re.compile(r"(\w+)\.(\w+) AS \"(.*?)\""),  # 通常のカラムマッピング
        re.compile(r"DATE_FORMAT\((\w+)\.(\w+),\s*'.*?'\)\s*AS \"(.*?)\""),  # DATE_FORMAT関数
        re.compile(r"CASE.*?END AS \"(.*?)\"")  # CASE文
    ]
    
    mapping = {}
    for pattern in patterns:
        matches = pattern.findall(sql_query)
        for match in matches:
            if len(match) == 3:  # DATE_FORMATや通常のカラム
                column_alias = match[2]
                column_full_name = f"{match[0]}.{match[1]}"
            elif len(match) == 1:  # CASE文
                column_alias = match[0]
                column_full_name = column_alias  # CASE文の場合、エイリアス名をそのまま使う
            mapping[column_alias] = column_full_name

    return mapping

# 元SQLファイル文に指定条件を挿入
def add_conditions_to_sql(sql_query, input_values, input_fields_types, deletion_exclusion, skip_deletion_exclusion=False):
    try:
        # サブクエリを検出して置換
        sql_query, subqueries = detect_and_replace_subqueries(sql_query)

        columns_mapping = extract_columns_mapping(sql_query)
        additional_conditions = []

        # その他の条件の生成
        for db_item, values in input_values.items():
            if db_item in columns_mapping and values:
                column_name = columns_mapping[db_item]
                if input_fields_types.get(db_item) == 'Date' and isinstance(values, dict):
                    start_date, end_date = values.get('start_date'), values.get('end_date')
                    if start_date and end_date:
                        if start_date == end_date:
                            condition = f"DATE({column_name}) = STR_TO_DATE('{start_date}', '%Y/%m/%d')"
                        else:
                            condition = f"{column_name} BETWEEN STR_TO_DATE('{start_date}', '%Y/%m/%d') AND STR_TO_DATE('{end_date}', '%Y/%m/%d')"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"日付条件: {condition}")
                elif input_fields_types.get(db_item) == 'FA' and isinstance(values, str) and values.strip():
                    condition = f"{column_name} LIKE '%{values}%'"
                    additional_conditions.append(condition)
                    LOGGER.debug(f"FA条件: {condition}")
                elif input_fields_types.get(db_item) == 'JSON' and isinstance(values, dict):
                    for path, value in values.items():
                        condition = f"JSON_CONTAINS({column_name}, '\"{value}\"', '$.{path}')"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"JSON条件 (dict): {condition}")
                elif input_fields_types.get(db_item) == 'JSON' and isinstance(values, str):
                    condition = f"JSON_CONTAINS({column_name}, '\"{values}\"', '$')"
                    additional_conditions.append(condition)
                    LOGGER.debug(f"JSON条件 (str): {condition}")
                elif values:
                    if isinstance(values, list):
                        placeholders = ', '.join([f"'{value}'" for value in values])
                        condition = f"{column_name} IN ({placeholders})"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"IN条件: {condition}")
                    else:
                        condition = f"{column_name} = '{values}'"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"等価条件: {condition}")

        # カラム名からテーブルエイリアスを特定
        table_alias = find_table_alias(sql_query)

        # 削除除外条件の追加
        if not skip_deletion_exclusion:
            deletion_exclusion = str(deletion_exclusion).upper()
            if deletion_exclusion == 'TRUE':
                additional_conditions.append(f"{table_alias}.deleted_at IS NULL")
                LOGGER.debug("削除除外条件が追加されました")

        # SQLクエリの前処理
        sql_query = preprocess_sql_query(sql_query)

        # GROUP BY句の検出と除去
        sql_query, group_by_clause = detect_and_remove_group_by(sql_query)

        # WHERE句の存在チェックと追加
        sql_query = check_and_prepare_where_clause(sql_query, additional_conditions)

        # GROUP BY句を再追加
        if group_by_clause:
            sql_query += "\n" + group_by_clause
        else:
            sql_query += ";"
        # サブクエリを元に戻す
        sql_query = restore_subqueries(sql_query, subqueries)
        return sql_query
    except Exception as e:
        LOGGER.error(f"add_conditions_to_sql関数内でエラーが発生しました: {e}")
        raise

def set_period_condition(period_condition, period_criteria, sql_query, category):
    """SQLクエリに期間条件を設定する

    Args:
        period_condition (str): 期間条件（例：'当日', '前日', '1か月前の月初未満'）
        period_criteria (str): 期間の基準（'登録日時', '更新日時', '応募：提出日時'）
        sql_query (str): 対象のSQLクエリ
        category (str): カテゴリ（'マスタ'の場合は期間条件を適用しない）

    Returns:
        str: 期間条件が適用されたSQLクエリ

    Raises:
        ValueError: 不正な期間条件が指定された場合
    """
    try:
        LOGGER.info("=" * 80)
        LOGGER.info("set_period_condition関数開始")
        LOGGER.info(f"引数 - period_condition: '{period_condition}'")
        LOGGER.info(f"引数 - period_criteria: '{period_criteria}'")
        LOGGER.info(f"引数 - category: '{category}'")
        LOGGER.info(f"引数 - sql_query長: {len(sql_query) if sql_query else 0}文字")
        
        # SQLクエリの前処理
        sql_query = preprocess_sql_query(sql_query)

        # ブランクまたは "マスタ" カテゴリの場合、そのままクエリを返す
        if not period_condition or category == 'マスタ':
            LOGGER.info("期間条件がブランク、またはカテゴリが 'マスタ' のため、クエリをそのまま返します。")
            return sql_query + ";"

        # GROUP BY句の検出と除去
        sql_query, group_by_clause = detect_and_remove_group_by(sql_query)

        # 基本のテーブルエイリアスを特定
        base_table_alias = find_table_alias(sql_query)
        LOGGER.info(f"基本テーブルエイリアス: '{base_table_alias}'")

        # 期間条件の設定
        LOGGER.info(f"期間基準の判定開始: '{period_criteria}'")
        if period_criteria == '登録日時':
            LOGGER.info("期間基準: 登録日時 - created_atを使用")
            column_name = 'created_at'
            table_alias = base_table_alias
        elif period_criteria == '更新日時':
            LOGGER.info("期間基準: 更新日時 - updated_atを使用")
            column_name = 'updated_at'
            table_alias = base_table_alias
            if not table_alias:
                raise ValueError("テーブルが見つかりません")
        elif period_criteria == 'ログイン日時':
            LOGGER.info("期間基準: ログイン日時 - last_login_atを使用")
            table_alias = base_table_alias
            if not table_alias:
                raise ValueError("テーブルが見つかりません")
                
            # usersテーブルの場合の処理
            if 'users' in sql_query.lower():
                # ログイン日時と更新日時の最新の日付を取得基準にする
                condition = generate_period_condition_for_login(period_condition, table_alias)
                if condition:
                    additional_conditions = [condition]
                    sql_query = check_and_prepare_where_clause(sql_query, additional_conditions)

                # GROUP BY句を再追加
                if group_by_clause:
                    sql_query += "\n" + group_by_clause
                else:
                    sql_query += ";"
                
                return sql_query
            else:
                column_name = 'last_login_at'
        elif period_criteria == '最終提出日時':
            LOGGER.info("期間基準: 最終提出日時 - last_submission_datetimeを使用")
            column_name = 'last_submission_datetime'
            # 最終提出日時は応募テーブル(user_applications)のuaエイリアスを使用
            LOGGER.info("user_applicationsテーブルのエイリアス検索開始")
            table_alias = find_submission_table_alias(sql_query)
            if not table_alias:
                # user_applicationsテーブルが見つからない場合は強制的に'ua'を使用
                table_alias = 'ua'
                LOGGER.warning("user_applicationsテーブルのエイリアスが検出できないため、'ua'を使用します")
            LOGGER.info(f"最終提出日時の期間条件で使用するエイリアス: '{table_alias}'")
        elif period_criteria == '提出期限':
            LOGGER.info("期間基準: 提出期限 - submission_deadlineを使用")
            column_name = 'submission_deadline'
            # 提出期限も応募テーブル(user_applications)のuaエイリアスを使用
            LOGGER.info("user_applicationsテーブルのエイリアス検索開始")
            table_alias = find_submission_table_alias(sql_query)
            if not table_alias:
                # user_applicationsテーブルが見つからない場合は強制的に'ua'を使用
                table_alias = 'ua'
                LOGGER.warning("user_applicationsテーブルのエイリアスが検出できないため、'ua'を使用します")
            LOGGER.info(f"提出期限の期間条件で使用するエイリアス: '{table_alias}'")
        else:
            LOGGER.error(f"不正な期間基準が指定されました: '{period_criteria}'")
            LOGGER.error(f"対応している期間基準: 申込日, 削除日, 最終ログイン, 提出期限")
            raise ValueError(f"不正な期間条件が指定されました: {period_criteria}")

        # 期間条件の生成と適用
        LOGGER.info(f"期間条件生成パラメータ: period_condition={period_condition}, column_name={column_name}, table_alias={table_alias}")
        condition = generate_period_condition(period_condition, column_name, table_alias)
        if condition:
            LOGGER.info(f"生成された期間条件: {condition}")
            additional_conditions = [condition]
            sql_query = check_and_prepare_where_clause(sql_query, additional_conditions)
        else:
            LOGGER.error("期間条件の生成に失敗しました")

        # GROUP BY句を再追加
        if group_by_clause:
            sql_query += "\n" + group_by_clause
            LOGGER.info("GROUP BY句を再追加しました")
        else:
            sql_query += ";"
            LOGGER.info("クエリ終端にセミコロンを追加しました")

        LOGGER.info("期間条件設定完了")
        LOGGER.info("=" * 50)
        # SQL本文のログ出力は抑制（セキュリティ/可読性のため）
        LOGGER.info("最終生成SQL（本文非表示）")
        LOGGER.info(f"SQL長: {len(sql_query)} 文字")
        LOGGER.info("=" * 50)
        
        return sql_query

    except Exception as e:
        LOGGER.error(f"set_period_condition関数内でエラーが発生しました: {e}")
        LOGGER.error(f"エラー詳細:\n{traceback.format_exc()}")
        raise

def generate_period_condition_for_login(period_condition, table_alias):
    """ログイン日時用の期間条件を生成する。更新日時とログイン日時の最新の日付を基準にする。"""
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    updated_at_column = f"{table_alias}.updated_at" if table_alias else "updated_at"
    last_login_at_column = f"{table_alias}.last_login_at" if table_alias else "last_login_at"
    
    if period_condition == '当日':
        return f"(DATE({updated_at_column}) = '{today}' OR DATE({last_login_at_column}) = '{today}')"
    elif period_condition == '前日':
        return f"(DATE({updated_at_column}) = '{yesterday}' OR DATE({last_login_at_column}) = '{yesterday}')"
    elif period_condition == '前日まで累積':
        return f"(DATE({updated_at_column}) <= '{yesterday}' OR DATE({last_login_at_column}) <= '{yesterday}')"
    elif period_condition == '当日まで累積':
        return f"(DATE({updated_at_column}) <= '{today}' OR DATE({last_login_at_column}) <= '{today}')"
    elif '～前日まで累積' in period_condition:
        start_date_str = period_condition.split('～')[0].strip()
        start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
        return f"(DATE({updated_at_column}) BETWEEN '{start_date}' AND '{yesterday}' OR DATE({last_login_at_column}) BETWEEN '{start_date}' AND '{yesterday}')"
    elif '～当日まで累積' in period_condition:
        start_date_str = period_condition.split('～')[0].strip()
        start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
        return f"(DATE({updated_at_column}) BETWEEN '{start_date}' AND '{today}' OR DATE({last_login_at_column}) BETWEEN '{start_date}' AND '{today}')"
    elif '～' in period_condition:
        parts = period_condition.split('～')
        if len(parts) == 3 and parts[2].strip() == 'までの期間':
            start_date_str, end_date_str = parts[0:2]
            start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
            end_date = datetime.strptime(end_date_str.strip(), '%Y年%m月%d日').date()
            return f"(DATE({updated_at_column}) BETWEEN '{start_date}' AND '{end_date}' OR DATE({last_login_at_column}) BETWEEN '{start_date}' AND '{end_date}')"
        elif 'まで累積' not in period_condition:
            start_date_str, end_date_str = parts
            start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
            end_date = datetime.strptime(end_date_str.strip(), '%Y年%m月%d日').date()
            return f"(DATE({updated_at_column}) BETWEEN '{start_date}' AND '{end_date}' OR DATE({last_login_at_column}) BETWEEN '{start_date}' AND '{end_date}')"
    elif '日前時点を1日分' in period_condition:
        days_ago = int(period_condition.split('日前')[0])
        target_date = today - timedelta(days=days_ago)
        return f"(DATE({updated_at_column}) = '{target_date}' OR DATE({last_login_at_column}) = '{target_date}')"
    elif '年' in period_condition and '月' in period_condition and '日' in period_condition:
        # 特定の日付（YYYY年MM月DD日）の場合
        specific_date = datetime.strptime(period_condition.strip(), '%Y年%m月%d日').date()
        return f"(DATE({updated_at_column}) = '{specific_date}' OR DATE({last_login_at_column}) = '{specific_date}')"
    else:
        LOGGER.warning(f"不明な期間条件: {period_condition}")
        return ""

def find_submission_table_alias(sql_query):
    """
    応募テーブルのエイリアスを検索する補助関数
    """
    LOGGER.info("find_submission_table_alias関数開始")
    LOGGER.info(f"SQL検索対象文字数: {len(sql_query)}")
    
    # 複数のパターンでuser_applicationsテーブルのエイリアスを検索
    patterns = [
        r'user_applications\s+(?:AS\s+)?([a-zA-Z][a-zA-Z0-9_]*)',  # user_applications ua
        r'FROM\s+user_applications\s+(?:AS\s+)?([a-zA-Z][a-zA-Z0-9_]*)',  # FROM user_applications ua
        r'JOIN\s+user_applications\s+(?:AS\s+)?([a-zA-Z][a-zA-Z0-9_]*)',  # JOIN user_applications ua
        r'user_applications\s+([a-zA-Z][a-zA-Z0-9_]*)',  # user_applications ua (ASなし)
    ]
    
    for i, pattern in enumerate(patterns):
        LOGGER.info(f"パターン{i+1}での検索: {pattern}")
        match = re.search(pattern, sql_query, re.IGNORECASE)
        if match:
            alias = match.group(1)
            LOGGER.info(f"パターン{i+1}でuser_applicationsテーブルのエイリアス検出: '{alias}'")
            return alias
        else:
            LOGGER.info(f"パターン{i+1}では検出されませんでした")
    
    # last_submission_datetimeが直接使用されている場合、そのエイリアスを探す
    LOGGER.info("last_submission_datetimeの直接使用からエイリアス検索")
    alias_pattern = r'([a-zA-Z][a-zA-Z0-9_]*)\.last_submission_datetime'
    alias_match = re.search(alias_pattern, sql_query, re.IGNORECASE)
    if alias_match:
        alias = alias_match.group(1)
        LOGGER.info(f"last_submission_datetimeの使用からエイリアス検出: '{alias}'")
        return alias
    else:
        LOGGER.info("last_submission_datetimeの直接使用も検出されませんでした")
    
    # SQLの一部を表示してデバッグ
    LOGGER.warning("user_applicationsテーブルのエイリアスが見つかりませんでした")
    # SQL本文のログ出力は抑制
    LOGGER.info(f"SQL長: {len(sql_query)} 文字（本文非表示）")
    
    return None

# テーブルエイリアスを特定する関数
def find_table_alias(sql_query):
    # '-- FROM clause' コメント以降の部分を抽出
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        LOGGER.warning("-- FROM clause not found in the SQL query.")
        return None  # コメントが見つからない場合はNoneを返す
    
    # FROM句以降のテキストを取得
    sub_query = sql_query[from_clause_index + len('-- FROM clause'):]

    # 実際のFROM句を探す。ASキーワードの存在にも対応。
    from_match = re.search(r'\bFROM\b\s+(\w+)\s+(?:AS\s+)?(\w+)', sub_query, re.IGNORECASE)
    if from_match:
        # テーブル名
        table_name = from_match.group(1)
        # エイリアスが指定されている場合はそれを使用し、なければテーブル名をそのままエイリアスとして使用
        alias = from_match.group(2) if from_match.group(2) else table_name
        return alias
    LOGGER.warning("FROM句のエイリアスが見つかりませんでした。")
    return None

# SQLクエリの前処理を行う。セミコロンの削除とトリミングを含む。
def preprocess_sql_query(sql_query):
    sql_query = sql_query.strip()
    if sql_query.endswith(";"):
        sql_query = sql_query[:-1]
    return sql_query.strip()

# SQLクエリからGROUP BY句を検出し、取り除いてそれを返す。
def detect_and_remove_group_by(sql_query):
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        return sql_query, ""
    
    sub_query = sql_query[from_clause_index:]
    upper_sub_query = sub_query.upper()
    group_by_index = upper_sub_query.find("GROUP BY")
    if group_by_index != -1:
        group_by_clause = sub_query[group_by_index:].strip()
        sub_query = sub_query[:group_by_index].strip()
        sql_query = sql_query[:from_clause_index] + sub_query
        return sql_query, group_by_clause
    return sql_query, ""

# サブクエリを検出
def detect_and_replace_subqueries(sql_query):
    subquery_pattern = r'-- subquery start(.*?)-- subquery end'
    subqueries = re.findall(subquery_pattern, sql_query, re.DOTALL)
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sql_query = sql_query.replace(f'-- subquery start{subquery}-- subquery end', placeholder)
    return sql_query, subqueries

# 元に戻す
def restore_subqueries(sql_query, subqueries):
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sql_query = sql_query.replace(placeholder, f'-- subquery start{subquery}-- subquery end')
    return sql_query

# SQLクエリにWHERE句があるか確認し、適切な形で返す。
def check_and_prepare_where_clause(sql_query, additional_conditions):
    """WHERE句の存在を確認し、必要に応じてANDを追加する。追加条件が存在しない場合は何も追加しない。 """
    if not additional_conditions:
        return sql_query  # 条件がなければそのまま返す

    # "-- FROM clause"以降の部分を取得
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        raise ValueError("-- FROM clause not found in the SQL query")

    # FROM句の後の部分を取得
    sub_query = sql_query[from_clause_index + len('-- FROM clause'):].strip()

    # サブクエリを一時的に置換
    subquery_pattern = r'-- subquery start(.*?)-- subquery end'
    subqueries = re.findall(subquery_pattern, sub_query, re.DOTALL)
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sub_query = sub_query.replace(f'-- subquery start{subquery}-- subquery end', placeholder)

    # WHERE句があるかどうかをチェック
    where_index = -1
    for match in re.finditer(r'\bWHERE\b', sub_query, re.IGNORECASE):
        if not re.search(r'__SUBQUERY_PLACEHOLDER_\d+__', sub_query[:match.start()]):
            where_index = match.start()
            break

    if where_index == -1:
        # WHERE句がない場合、WHERE句を追加
        modified_sub_query = sub_query + "\nWHERE " + " AND ".join(additional_conditions)
    else:
        # WHERE句がある場合、ANDで条件を追加
        where_clause = sub_query[:where_index + len("WHERE")]
        remaining_query = sub_query[where_index + len("WHERE"):].strip()
        modified_sub_query = where_clause + " " + remaining_query + " AND " + " AND ".join(additional_conditions)

    # サブクエリを元に戻す
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        modified_sub_query = modified_sub_query.replace(placeholder, f'-- subquery start{subquery}-- subquery end')

    # 元のクエリのFROM句までの部分と、修正後のサブクエリを結合
    final_query = sql_query[:from_clause_index + len('-- FROM clause')] + "\n" + modified_sub_query
    return final_query

def generate_period_condition(period_condition, column_name, table_alias):
    """期間条件に基づくWHERE句の条件を生成する。"""
    try:
        LOGGER.info(f"期間条件生成開始: period_condition='{period_condition}', column_name='{column_name}', table_alias='{table_alias}'")
        
        if not period_condition or not period_condition.strip():
            LOGGER.warning("期間条件が空です")
            return ""
        
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)

        date_column = f"{table_alias}.{column_name}" if table_alias else column_name
        LOGGER.info(f"期間条件生成中: date_column={date_column}, period_condition={period_condition}")
    except Exception as e:
        LOGGER.error(f"期間条件生成の初期化エラー: {e}")
        return ""
    
    try:
        if period_condition == '月初未満':
            # 現在の月の月初を計算
            current_month_start = date(today.year, today.month, 1)
            condition = f" DATE({date_column}) < '{current_month_start}'"
            LOGGER.info(f"月初未満条件生成: {condition}")
            return condition
        
        elif period_condition == '当日':
            condition = f" DATE({date_column}) = '{today}'"
            LOGGER.info(f"当日条件生成: {condition}")
            return condition
        elif period_condition == '前日':
            condition = f" DATE({date_column}) = '{yesterday}'"
            LOGGER.info(f"前日条件生成: {condition}")
            return condition
        elif period_condition == '前日まで累積':
            condition = f" DATE({date_column}) <= '{yesterday}'"
            LOGGER.info(f"前日まで累積条件生成: {condition}")
            return condition
        elif period_condition == '当日まで累積':
            condition = f" DATE({date_column}) <= '{today}'"
            LOGGER.info(f"当日まで累積条件生成: {condition}")
            return condition
        elif '～前日まで累積' in period_condition:
            start_date_str = period_condition.split('～')[0].strip()
            start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
            condition = f" DATE({date_column}) BETWEEN '{start_date}' AND '{yesterday}'"
            LOGGER.info(f"前日まで累積範囲条件生成: {condition}")
            return condition
        elif '～当日まで累積' in period_condition:
            start_date_str = period_condition.split('～')[0].strip()
            start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
            condition = f" DATE({date_column}) BETWEEN '{start_date}' AND '{today}'"
            LOGGER.info(f"当日まで累積範囲条件生成: {condition}")
            return condition
        elif '～' in period_condition:
            parts = period_condition.split('～')
            if len(parts) == 3 and parts[2].strip() == 'までの期間':
                start_date_str, end_date_str = parts[0:2]
                start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
                end_date = datetime.strptime(end_date_str.strip(), '%Y年%m月%d日').date()
                condition = f" DATE({date_column}) BETWEEN '{start_date}' AND '{end_date}'"
                LOGGER.info(f"期間範囲条件生成: {condition}")
                return condition
            elif 'まで累積' not in period_condition:
                try:
                    start_date_str, end_date_str = parts
                    LOGGER.info(f"期間条件分割結果: 開始='{start_date_str}', 終了='{end_date_str}'")
                    
                    # 開始日の解析
                    start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
                    LOGGER.info(f"開始日解析成功: {start_date}")
                    
                    # 終了日の処理：「まで」「までの期間」等の文字を除去
                    original_end_date_str = end_date_str
                    end_date_str = end_date_str.strip()
                    end_date_str = end_date_str.replace('まで', '').replace('の期間', '').strip()
                    
                    # 年が省略されている場合は開始日の年を使用
                    if '年' not in end_date_str:
                        start_year = start_date.year
                        end_date_str = f"{start_year}年{end_date_str}"
                    
                    LOGGER.info(f"期間条件解析: 開始日='{start_date_str}' -> {start_date}, 終了日='{original_end_date_str}' -> '{end_date_str}'")
                    
                    try:
                        end_date = datetime.strptime(end_date_str, '%Y年%m月%d日').date()
                        condition = f" DATE({date_column}) BETWEEN '{start_date}' AND '{end_date}'"
                        LOGGER.info(f"生成された期間条件: {condition}")
                        return condition
                    except ValueError as e:
                        LOGGER.error(f"終了日の解析エラー: '{end_date_str}' - {e}")
                        # フォールバック: 異なる形式を試行
                        try:
                            # 「月日」形式を試行
                            if '月' in end_date_str and '日' in end_date_str:
                                end_date_str_fallback = end_date_str.replace('月', '/').replace('日', '')
                                end_date = datetime.strptime(f"{start_date.year}/{end_date_str_fallback}", '%Y/%m/%d').date()
                                condition = f" DATE({date_column}) BETWEEN '{start_date}' AND '{end_date}'"
                                LOGGER.info(f"フォールバック成功 - 生成された期間条件: {condition}")
                                return condition
                        except ValueError as fallback_e:
                            LOGGER.error(f"フォールバック解析も失敗: {fallback_e}")
                        return ""
                except Exception as e:
                    LOGGER.error(f"期間条件解析全体エラー: {e}")
                    LOGGER.error(f"エラー詳細: parts={parts}")
                    return ""
        elif '日前時点を1日分' in period_condition:
            days_ago = int(period_condition.split('日前')[0])
            target_date = today - timedelta(days=days_ago)
            condition = f" DATE({date_column}) = '{target_date}'"
            LOGGER.info(f"日前時点条件生成: {condition}")
            return condition
        elif '年' in period_condition and '月' in period_condition and '日' in period_condition:
            # 特定の日付（YYYY年MM月DD日）の場合
            specific_date = datetime.strptime(period_condition.strip(), '%Y年%m月%d日').date()
            condition = f" DATE({date_column}) = '{specific_date}'"
            LOGGER.info(f"特定日付条件生成: {condition}")
            return condition
        else:
            LOGGER.warning(f"不明な期間条件: '{period_condition}'")
            return ""
    except Exception as e:
        LOGGER.error(f"期間条件生成処理エラー: {e}")
        LOGGER.error(f"period_condition: '{period_condition}'")
        return ""

def get_column_letter(column_index):
    # 1から始まるインデックスに対応するため、1を引く
    column_index -= 1
    letter = ''
    while column_index >= 0:
        letter = chr(column_index % 26 + 65) + letter
        column_index = column_index // 26 - 1
    return letter

# スプシ貼り付け
@retry_on_exception
def export_to_spreadsheet(conn, sql_query, save_path_id, sheet_name, json_keyfile_path, paste_format, main_sheet_name, csv_file_name_column, main_table_name, category, chunk_size=10000, delay=0):

    LOGGER.info("SQLクエリの実行を開始します。（スプシの貼り付け）")
    cursor = conn.cursor()
    try:
        cursor.execute(sql_query)
        fetchall = cursor.fetchall()
        record_count = len(fetchall)

        # ヘッダ行の取得
        headers = [i[0] for i in cursor.description]

        # Decimal、date、datetimeオブジェクトを文字列に変換
        converted_data = []
        LOGGER.info(f"データ変換開始: 総レコード数 {record_count}")
        for row_idx, row in enumerate(fetchall):
            converted_row = []
            for col_idx, cell in enumerate(row):
                if isinstance(cell, (Decimal, date, datetime, timedelta)):
                    converted_cell = str(cell)
                    # サンプル行のみ詳細ログ出力（最初の10行のみ）
                    if row_idx < 10:
                        LOGGER.debug(f"行{row_idx+1}, 列{col_idx+1}: {type(cell).__name__} -> str: {converted_cell}")
                elif cell is not None:
                    converted_cell = cell
                else:
                    converted_cell = ''
                converted_row.append(converted_cell)
            converted_data.append(converted_row)
        
        # データ変換後の型チェック
        LOGGER.info("データ変換完了。型チェックを実行します。")
        for row_idx, row in enumerate(converted_data[:3]):  # 最初の3行をサンプルチェック
            for col_idx, cell in enumerate(row):
                cell_type = type(cell).__name__
                if cell_type not in ['str', 'int', 'float', 'bool', 'NoneType']:
                    LOGGER.warning(f"想定外の型が検出されました - 行{row_idx+1}, 列{col_idx+1}: {cell_type} = {cell}")
                else:
                    LOGGER.debug(f"行{row_idx+1}, 列{col_idx+1}: {cell_type} = {str(cell)[:50]}")  # 50文字まで表示
        
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "成功", None, save_path_id)  # ログシートに成功を書き込む

        SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = authenticate_google_api(json_keyfile_path, SCOPES)
        gc = gspread.authorize(credentials)

        spreadsheet = gc.open_by_key(save_path_id)

        if sheet_name == main_sheet_name:
            # 実行シートの場合は処理をスキップ
            LOGGER.info(f"Skipping export to main sheet: {sheet_name}")
            return

        if not sheet_name:
            # "CSVファイル名/SSシート名" がブランクの場合、"CSVファイル呼称" 列の値を使用
            sheet_name = csv_file_name_column
            LOGGER.info("CSVファイル名/SSシート名がブランクの場合、CSVファイル呼称を使用します。")

        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            # 指定されたシートが存在しない場合は新しいシートを追加
            LOGGER.info(f"シート '{sheet_name}' が存在しないため、新規作成します。")
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=26)
            # 新しいシートの場合、ヘッダ行を追加し、ブランク行を追加
            worksheet.update('A1', [headers])
            worksheet.append_row([''] * len(headers))  # ブランク行の追加

        # 既存シートの1行目をチェックし、ヘッダ行が存在しない場合は追加
        if not worksheet.row_values(1):
            worksheet.update('A1', [headers])
            worksheet.append_row([''] * len(headers))  # ブランク行の追加
            LOGGER.info("ヘッダ行をシートに追加しました。")
        else:
            LOGGER.info("ヘッダ行は既に存在します。")

        column_count = len(cursor.description)
        last_column_letter = get_column_letter(column_count)
        LOGGER.info(f"Column count: {column_count}, Last column letter: {last_column_letter}")

        if paste_format == '最終行積立て':
            last_row = len(worksheet.col_values(1)) + 1
            if last_row > worksheet.row_count:
                additional_rows = last_row - worksheet.row_count
                worksheet.add_rows(additional_rows)
            for i in range(0, len(converted_data), chunk_size):
                chunk = converted_data[i:i + chunk_size]
                LOGGER.info(f"チャンク{i//chunk_size + 1} (行{i+1}～{min(i+chunk_size, len(converted_data))}) の書き込みを開始")
                
                # チャンクの型チェック
                for chunk_row_idx, chunk_row in enumerate(chunk[:1]):  # 最初の行のみチェック
                    for col_idx, cell in enumerate(chunk_row):
                        cell_type = type(cell).__name__
                        if cell_type not in ['str', 'int', 'float', 'bool', 'NoneType']:
                            LOGGER.error(f"チャンク内で想定外の型が検出: 行{chunk_row_idx+1}, 列{col_idx+1}: {cell_type} = {cell}")
                
                try:
                    worksheet.update(f'A{last_row}', chunk)
                    LOGGER.info(f"チャンク{i//chunk_size + 1} の書き込み完了")
                except Exception as chunk_error:
                    LOGGER.error(f"チャンク{i//chunk_size + 1} の書き込み中にエラー: {chunk_error}")
                    LOGGER.error(f"エラー詳細:\n{traceback.format_exc()}")
                    raise
                    
                last_row += len(chunk)
                if delay:
                    LOGGER.info(f"{delay}秒待機します。")
                    time.sleep(delay)
        elif paste_format == '全張替え':
            # 既存のヘッダ行を保持し、データ部分のみクリア
            clear_range = f'A2:{last_column_letter}{worksheet.row_count}'
            LOGGER.info(f"Clearing range: {clear_range}")
            worksheet.batch_clear([clear_range])

            data_row_count = len(converted_data)
            rows_to_add = data_row_count - (worksheet.row_count - 1)
            # 1万行ずつ追加
            while rows_to_add > 0:
                rows_to_add_now = min(10000, rows_to_add)
                worksheet.add_rows(rows_to_add_now)
                rows_to_add -= rows_to_add_now
                LOGGER.info(f"{rows_to_add_now} 行をシートに追加しました。")

            row_start = 2  # データの開始行は2行目から
            for i in range(0, len(converted_data), chunk_size):
                chunk = converted_data[i:i + chunk_size]
                LOGGER.info(f"全張替えチャンク{i//chunk_size + 1} (行{i+1}～{min(i+chunk_size, len(converted_data))}) の書き込みを開始")
                
                # チャンクの型チェック
                for chunk_row_idx, chunk_row in enumerate(chunk[:1]):  # 最初の行のみチェック
                    for col_idx, cell in enumerate(chunk_row):
                        cell_type = type(cell).__name__
                        if cell_type not in ['str', 'int', 'float', 'bool', 'NoneType']:
                            LOGGER.error(f"全張替えチャンク内で想定外の型が検出: 行{chunk_row_idx+1}, 列{col_idx+1}: {cell_type} = {cell}")
                
                try:
                    worksheet.update(f'A{row_start}', chunk)
                    LOGGER.info(f"全張替えチャンク{i//chunk_size + 1} の書き込み完了")
                except Exception as chunk_error:
                    LOGGER.error(f"全張替えチャンク{i//chunk_size + 1} の書き込み中にエラー: {chunk_error}")
                    LOGGER.error(f"エラー詳細:\n{traceback.format_exc()}")
                    raise
                    
                row_start += len(chunk)
                if delay:
                    LOGGER.info(f"{delay}秒待機します。")
                    time.sleep(delay)

        LOGGER.info(f"Data has been transferred to {sheet_name} sheet in {save_path_id} with {paste_format} method.")
    except Exception as e:
        LOGGER.error(f"Error during query execution or spreadsheet writing: {e}")
        LOGGER.error(f"エラー発生時の状況:")
        LOGGER.error(f"  - sheet_name: {sheet_name}")
        LOGGER.error(f"  - paste_format: {paste_format}")
        LOGGER.error(f"  - record_count: {record_count}")
        LOGGER.error(f"  - chunk_size: {chunk_size}")
        LOGGER.error(f"完全なスタックトレース:\n{traceback.format_exc()}")
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "失敗", str(e), save_path_id)  # ログシートに失敗を書き込む
        raise
    finally:
        cursor.close()

# テスト実行
@retry_on_exception
def setup_test_environment(test_execution, output_to_spreadsheet, save_path_id, csv_file_name, spreadsheet_id, json_keyfile_path):
    if str(test_execution).lower() == 'true':
        if output_to_spreadsheet == 'CSV':
            test_folder = os.path.join(save_path_id, 'test')
            if not os.path.exists(test_folder):
                os.makedirs(test_folder)
                LOGGER.info(f"テスト用フォルダを作成しました: {test_folder}")
            save_path_id = test_folder
        elif output_to_spreadsheet == 'スプシ':
            SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            credentials = authenticate_google_api(json_keyfile_path, SCOPES)
            gc = gspread.authorize(credentials)
            
            spreadsheet_id = save_path_id
            
            try:
                spreadsheet = gc.open_by_key(save_path_id)
                test_sheet_name = f"{csv_file_name}_test"
                try:
                    worksheet = spreadsheet.worksheet(test_sheet_name)
                except gspread.exceptions.WorksheetNotFound:
                    try:
                        source_sheet = spreadsheet.worksheet(csv_file_name)
                        worksheet = spreadsheet.add_worksheet(title=test_sheet_name, rows=source_sheet.row_count, cols=source_sheet.col_count)
                        header_row = source_sheet.row_values(1)
                        worksheet.insert_row(header_row, index=1)
                        LOGGER.info(f"テスト用シート '{test_sheet_name}' を作成しました。")
                    except gspread.exceptions.WorksheetNotFound:
                        raise ValueError(f"The source sheet '{csv_file_name}' does not exist in the spreadsheet.")
                csv_file_name = test_sheet_name
            except Exception as e:
                LOGGER.error(f"テスト環境のセットアップ中にエラーが発生しました: {e}")
                raise

    return save_path_id, csv_file_name

# ログの書き出し
def write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, result, error_log=None, save_path_id=None):
    log_spreadsheet_id = '1iqDqeGXAovNQfnuuOi2xLzJIrmXOE1FKOgrSLgG0SOw'
    log_sheet_name = 'log'

    SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = authenticate_google_api(json_keyfile_path, SCOPES)
    client = gspread.authorize(credentials)

    try:
        spreadsheet = client.open_by_key(log_spreadsheet_id)
        worksheet = spreadsheet.worksheet(log_sheet_name)

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            csv_file_name_column, sheet_name, main_table_name, category, 
            save_path_id, record_count, result, error_log, timestamp, ''
        ]
        
        # ログデータの型チェック（簡素化）
        for idx, item in enumerate(row_data):
            item_type = type(item).__name__
            if item_type not in ['str', 'int', 'float', 'bool', 'NoneType']:
                LOGGER.warning(f"ログデータで想定外の型が検出: 項目{idx+1}: {item_type}")
        
        try:
            worksheet.append_row(row_data)
        except Exception as log_error:
            LOGGER.error(f"ログシートへの書き込み中にエラーが発生: {log_error}")
            LOGGER.error(f"ログ書き込みエラーの詳細:\n{traceback.format_exc()}")
            raise

        LOGGER.info(f"ログシートに書き込みました: {row_data}")
    except gspread.exceptions.APIError as e:
        LOGGER.error(f"ログシートへの書き込み中にエラーが発生しました: {e}")

# DB項目のデータ型を強制指定
def get_data_types(worksheet):
    headers = worksheet.row_values(1)
    data_types = {}

    db_item_col = None
    data_type_col = None

    for i, header in enumerate(headers):
        if header == 'DB項目':
            db_item_col = i
        elif header == 'DATA_TYPE':
            data_type_col = i

    if db_item_col is not None and data_type_col is not None:
        last_row = len(worksheet.col_values(db_item_col + 1))

        db_item_range = worksheet.range(2, db_item_col + 1, last_row, db_item_col + 1)
        data_type_range = worksheet.range(2, data_type_col + 1, last_row, data_type_col + 1)

        db_items = [cell.value for cell in db_item_range]
        data_types_list = [cell.value for cell in data_type_range]

        data_types = dict(zip(db_items, data_types_list))
    
    #LOGGER.info(f"取得したデータ型: {data_types}")
    return data_types

# Parquet用のデータ型を適用する関数（安全な変換）
def apply_data_types_to_df_for_parquet(df, data_types, LOGGER):
    converted_columns = []
    for column, data_type in data_types.items():
        if column in df.columns:
            try:
                if data_type == 'txt':
                    # Parquet用に文字列型を安全に変換
                    df[column] = df[column].astype(str)
                    converted_columns.append(column)
                elif data_type == 'int':
                    df[column] = pd.to_numeric(df[column], errors='coerce').astype('Int64')
                    converted_columns.append(column)
                elif data_type == 'float':
                    df[column] = pd.to_numeric(df[column], errors='coerce').astype('float64')
                    converted_columns.append(column)
                elif data_type == 'date':
                    df[column] = df[column].astype(str)
                    converted_columns.append(column)
                elif data_type == 'datetime':
                    df[column] = df[column].astype(str)
                    converted_columns.append(column)
            except Exception as e:
                LOGGER.warning(f"Column '{column}' conversion to '{data_type}' failed, keeping as string: {e}")
                df[column] = df[column].astype(str)
                converted_columns.append(column)
    
    if converted_columns:
        LOGGER.info(f"Parquet用型変換を実行しました: {converted_columns}")
    else:
        LOGGER.info("Parquet用型変換は行われませんでした。")
    
    return df

# データ型を適用する関数（フォーマット済み）
def apply_data_types_to_df(df, data_types, LOGGER, encoding='utf-8'):
    converted_columns = []  # 型変換を行った列名を格納するリスト
    for column, data_type in data_types.items():
        if column in df.columns:
            try:
                if data_type == 'txt':
                    df[column] = df[column].apply(lambda x: x.encode(encoding).decode(encoding) if isinstance(x, str) else x)
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'int':
                    df[column] = pd.to_numeric(df[column], errors='raise').astype('Int64')
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'float':
                    df[column] = pd.to_numeric(df[column], errors='raise').astype(float)
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'date':
                    df[column] = df[column].astype(str)  # date型は文字列に変換済み
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'datetime':
                    df[column] = df[column].astype(str)  # datetime型は文字列に変換済み
                    converted_columns.append(column)  # 型変換を行った列名を追加
            except ValueError as e:
                LOGGER.error(f"Error converting column '{column}' to type '{data_type}': {e}")
                raise ValueError(f"Error converting column '{column}' to type '{data_type}': {e}")
            except Exception as e:
                LOGGER.error(f"Unexpected error processing column '{column}': {e}")
                raise Exception(f"Unexpected error processing column '{column}': {e}")
    
    if converted_columns:
        LOGGER.info(f"以下の列の型変換を行いました: {', '.join(converted_columns)}")
    else:
        LOGGER.info("型変換は行われませんでした。")
    
    return df

# スプシに基づき編集するSQL
def execute_sql_query_with_conditions(sql_file_name, config, period_condition, period_criteria, deletion_exclusion, category, main_table_name):
    sql_query = load_sql_from_file(sql_file_name, config['google_folder_id'], config['json_keyfile_path'])
    if sql_query:
        try:
            input_values, input_fields_types = {}, {}
            sql_query_with_period_condition = set_period_condition(period_condition, period_criteria, sql_query, category)
            if category != 'マスタ':
                sql_query_with_conditions = add_conditions_to_sql(
                    sql_query_with_period_condition, input_values, input_fields_types, deletion_exclusion
                )
            else:
                sql_query_with_conditions = sql_query_with_period_condition
                # SQL本文のログ出力は抑制
                LOGGER.info(f"実行SQLクエリ - ファイル名: {sql_file_name}（本文非表示, 長さ: {len(sql_query_with_conditions)} 文字）")
            return sql_query_with_conditions
        except Exception as e:
            LOGGER.error(f"SQLクエリの処理中にエラーが発生しました: {e}")
            return None
    else:
        LOGGER.warning(f"{sql_file_name} の読み込みに失敗しました。代わりに 'SELECT *' を実行します。")
        return f"SELECT * -- FROM clause\nFROM {main_table_name}"
