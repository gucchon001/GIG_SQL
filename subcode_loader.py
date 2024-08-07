from oauth2client.service_account import ServiceAccountCredentials
import gspread
from googleapiclient.discovery import build
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import os
import re
import io
import time
from decimal import Decimal
from my_logging import setup_department_logger
from tenacity import retry, stop_after_attempt, wait_exponential, before_sleep_log
import shutil
import traceback

LOGGER = setup_department_logger('main')

#Googleの認証処理
def authenticate_google_api(json_keyfile_path, scopes):
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scopes)
    return credentials

#リトライ処理
def retry_on_exception(func):
    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=60, min=60, max=300),
        before_sleep=before_sleep_log
    )
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)
    return wrapper

def before_sleep_log(retry_state):
    LOGGER.warning(f"Retrying {retry_state.fn.__name__} after exception: {retry_state.outcome.exception()}")
    print(f"Retrying {retry_state.fn.__name__} after exception: {retry_state.outcome.exception()}")

@retry_on_exception
def load_sql_file_list_from_spreadsheet(spreadsheet_id, sheet_name, json_keyfile_path, execution_column):
    """
    指定されたGoogleスプレッドシートからSQLファイルのリストを読み込む関数

    :param spreadsheet_id: スプレッドシートのID。
    :param sheet_name: 読み込むシートの名前。
    :param json_keyfile_path: Google APIの認証情報が含まれるJSONファイルのパス。
    :param execution_column: 実行対象の列名。
    :return: 実行対象とマークされたSQLファイル名のリスト。
    """
    SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    SQL_FILE_COLUMN = 'sqlファイル名'
    CSV_FILE_COLUMN = 'CSVファイル名/SSシート名'
    FILENAME_FORMAT_COLUMN = '保存ファイル名形式'
    PERIOD_CONDITION_COLUMN = '取得期間'
    PERIOD_CRITERIA_COLUMN = '取得基準'
    SPREADSHEET_COLUMN = '出力先'
    SAVE_PATH_COLUMN = '保存先PATH/ID'
    DELETION_EXCLUSION_COLUMN = '削除R除外'
    PASTE_FORMAT_COLUMN = 'スプシ貼り付け形式'
    TEST_COLUMN = 'テスト'
    CATEGORY_COLUMN = 'カテゴリ'
    MAIN_TABLE_COLUMN = 'メインテーブル'
    CSV_FILE_NAME_COLUMN = 'CSVファイル呼称'
    SHEET_NAME_COLUMN = 'シート名' 

    SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = authenticate_google_api(json_keyfile_path, SCOPES)
    gc = gspread.authorize(credentials)

    worksheet = gc.open_by_key(spreadsheet_id).worksheet(sheet_name)
    records = worksheet.get_all_records()

    sql_and_csv_files = []
    for record in records:
        if record[execution_column] == 'TRUE':
            sql_file_name = record[SQL_FILE_COLUMN]
            csv_file_name = record[CSV_FILE_COLUMN]
            filename_format = record[FILENAME_FORMAT_COLUMN]
            period_condition = record[PERIOD_CONDITION_COLUMN]
            period_criteria = record[PERIOD_CRITERIA_COLUMN]
            save_path_id = record[SAVE_PATH_COLUMN]
            output_to_spreadsheet = record[SPREADSHEET_COLUMN]
            deletion_exclusion = record[DELETION_EXCLUSION_COLUMN]
            paste_format = record[PASTE_FORMAT_COLUMN]
            test_execution = record[TEST_COLUMN]
            category = record[CATEGORY_COLUMN]
            main_table_name = record[MAIN_TABLE_COLUMN]
            csv_file_name_column = record[CSV_FILE_NAME_COLUMN]
            sheet_name = record[SHEET_NAME_COLUMN]

            if filename_format:
                now = datetime.now()
                filename_without_ext = os.path.splitext(csv_file_name)[0]  # 拡張子を除いたファイル名
                filename_patterns = {
                    'yyyyMMdd_{filename}': now.strftime('%Y%m%d_') + filename_without_ext,
                    '{filename}_yyyy-MM-dd': filename_without_ext + '_' + now.strftime('%Y-%m-%d'),
                    '{filename}yyyy-MM-dd': filename_without_ext + now.strftime('%Y-%m-%d'),
                    '{filename}_yyyyMMddhhmmss': filename_without_ext + '_' + now.strftime('%Y%m%d%H%M%S'),
                    '{filename}_yyyy-MM': filename_without_ext + '_' + now.strftime('%Y-%m')
                }
                for pattern, value in filename_patterns.items():
                    if pattern in filename_format:
                        csv_file_name = filename_format.replace(pattern, value) + '.csv'
                        break
                else:
                    csv_file_name = filename_format + '.csv'
            sql_and_csv_files.append((sql_file_name, csv_file_name, period_condition, period_criteria, save_path_id, output_to_spreadsheet, deletion_exclusion, paste_format, test_execution, category, main_table_name, csv_file_name_column, sheet_name))

            print(f"Loaded sheet: {sheet_name}")
            print(f"Execution column: {execution_column}")
            print(f"Filtered data: {sql_and_csv_files}")
    return sql_and_csv_files

@retry_on_exception
def load_sql_from_file(file_path, google_folder_id, json_keyfile_path):
    try:
        SCOPES = ['https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, SCOPES)
        service = build('drive', 'v3', credentials=credentials)

        LOGGER.info(f"GoogleドライブのフォルダID: {google_folder_id}")
        LOGGER.info(f"SQLファイル名: {file_path}")

        query = f"'{google_folder_id}' in parents and name = '{file_path}'"
        file_list_results = service.files().list(q=query, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        files = file_list_results.get('files', [])

        LOGGER.info(f"ファイルリスト結果: {files}")

        if files:
            file_id = files[0]['id']
            file_content = service.files().get_media(fileId=file_id, supportsAllDrives=True).execute().decode('utf-8')
            sql_query = file_content.strip()
            #logger.info(f"SQLクエリ内容: {sql_query}")
            return sql_query
        else:
            error_message = f"ファイルが見つかりません: {file_path}"
            LOGGER.error(error_message)
            return None
    except Exception as e:
        error_message = f"SQLファイルの読み込み中にエラーが発生しました: {e}"
        LOGGER.error(error_message)
        LOGGER.error(traceback.format_exc())
        return None

# CSVファイルに保存する処理
def save_chunk_to_csv(chunk, file_path, include_header=True):
    mode = 'w' if include_header else 'a'  # ヘッダを含める場合は'w'、含めない場合は'a'
    header = include_header
    with open(file_path, mode=mode, newline='', encoding='cp932', errors='replace') as file:
        chunk.to_csv(file, index=False, header=header)

# データフレームをチャンクに分割して処理する関数
def process_dataframe_in_chunks(df, chunk_size, file_path, delay=None):
    if chunk_size is None or len(df) <= chunk_size:
        try:
            save_chunk_to_csv(df, file_path)
            return len(df)
        except Exception as e:
            print(f"ファイル書き込み時にエラーが発生しました: {e}")
            raise
    else:
        chunk_dir = os.path.join(os.path.dirname(file_path), 'chunk')
        os.makedirs(chunk_dir, exist_ok=True)
        chunk_file_paths = []
        total_records = 0
        for i, chunk in enumerate(pd.read_csv(io.StringIO(df.to_csv(index=False, encoding='cp932', sep=',')), chunksize=chunk_size, dtype=str)):
            chunk_file_path = os.path.join(chunk_dir, f"{os.path.basename(file_path)[:-4]}_chunk_{i}.csv")
            try:
                save_chunk_to_csv(chunk, chunk_file_path)
                chunk_file_paths.append(chunk_file_path)
                chunk_records = len(chunk)
                total_records += chunk_records
                if delay:
                    time.sleep(delay)
            except Exception as e:
                print(f"チャンクファイル書き込み時にエラーが発生しました: {e}")
                raise

        combine_chunk_files(chunk_file_paths, file_path)
        shutil.rmtree(chunk_dir)
        
        return total_records

def combine_chunk_files(chunk_file_paths, output_file_path):
    with open(output_file_path, 'w', encoding='cp932', errors='replace', newline='') as f_out:
        with open(chunk_file_paths[0], 'r', encoding='cp932', errors='replace') as f_in:
            lines = f_in.readlines()
            f_out.writelines(lines)  # ヘッダ行とデータ行を書き込む
        for file_path in chunk_file_paths[1:]:
            with open(file_path, 'r', encoding='cp932', errors='replace') as f_in:
                lines = f_in.readlines()
                f_out.writelines(lines[1:])  # ヘッダ行をスキップしてデータ行を書き込む

#CSVファイル処理
@retry_on_exception
def csvfile_export(conn, sql_query, csv_file_path, main_table_name, category, json_keyfile_path, spreadsheet_id, csv_file_name, csv_file_name_column, sheet_name, chunk_size=None, delay=None):
    try:
        if sheet_name:
            try:
                scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scope)
                gc = gspread.authorize(credentials)
                worksheet = gc.open_by_key(spreadsheet_id).worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound as e:
                LOGGER.error(f"ワークシート '{sheet_name}' が見つかりませんでした: {e}")
                write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
                raise

        data_types = get_data_types(worksheet) if sheet_name else {}
        df = pd.read_sql(sql_query, conn)

        # NaN、None、'nan'、'None'を空文字列に置換
        df = df.fillna('').replace({'None': '', 'nan': ''})

        # Int64型のカラムで空文字列を0に置換
        int64_columns = [col for col, dtype in data_types.items() if dtype == 'int' and col in df.columns]
        for col in int64_columns:
            df[col] = df[col].replace('', 0)

        # データ型を適用
        df = apply_data_types_to_df(df, data_types, LOGGER)

        # 数値型の列で空文字列になっているセルを0に置換（Int64型以外の数値型カラム用）
        numeric_columns = df.select_dtypes(include=['float64']).columns
        for col in numeric_columns:
            df[col] = df[col].replace('', 0)

        # データフレームの各要素を文字列に変換
        df = df.applymap(lambda x: str(int(x)) if isinstance(x, (float, Decimal)) and x.is_integer() else str(x) if not pd.isna(x) else '')

        # 一時ファイルパスを作成
        temp_file_path = csv_file_path + '.temp'

        try:
            record_count = process_dataframe_in_chunks(df, chunk_size, temp_file_path, delay=delay)
            
            # 処理が成功したら、一時ファイルを正式なファイルに置き換え
            if os.path.exists(csv_file_path):
                os.remove(csv_file_path)
            os.rename(temp_file_path, csv_file_path)

            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "成功", None, csv_file_path)
        except Exception as e:
            LOGGER.error(f"CSVファイル書き込み時にエラーが発生しました: {e}")
            write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
            
            # エラーが発生した場合、一時ファイルを削除
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            
            raise

    except Exception as e:
        LOGGER.error(f"クエリ実行またはCSVファイル書き込み時にエラーが発生しました: {e}")
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, 0, json_keyfile_path, "失敗", str(e), csv_file_path)
        raise

# 複数のパターンにマッチする正規表現を定義
def extract_columns_mapping(sql_query):
    patterns = [
        re.compile(r"(\w+)\.(\w+) AS \"(.*?)\""),  # 通常のカラムマッピング
        re.compile(r"DATE_FORMAT\((\w+)\.(\w+),\s*'.*?'\)\s*AS \"(.*?)\""),  # DATE_FORMAT関数
        re.compile(r"CASE.*?END AS \"(.*?)\"")  # CASE文
    ]
    
    mapping = {}
    for pattern in patterns:
        matches = pattern.findall(sql_query)
        for match in matches:
            if len(match) == 3:  # DATE_FORMATや通常のカラム
                column_alias = match[2]
                column_full_name = f"{match[0]}.{match[1]}"
            elif len(match) == 1:  # CASE文
                column_alias = match[0]
                column_full_name = column_alias  # CASE文の場合、エイリアス名をそのまま使う
            mapping[column_alias] = column_full_name

    return mapping

#元SQLファイル文に指定条件を挿入
def add_conditions_to_sql(sql_query, input_values, input_fields_types, deletion_exclusion, skip_deletion_exclusion=False):
    try:
        # サブクエリを検出して置換
        sql_query, subqueries = detect_and_replace_subqueries(sql_query)

        columns_mapping = extract_columns_mapping(sql_query)
        additional_conditions = []

        # その他の条件の生成
        for db_item, values in input_values.items():
            if db_item in columns_mapping and values:
                column_name = columns_mapping[db_item]
                if input_fields_types[db_item] == 'Date' and isinstance(values, dict):
                    start_date, end_date = values.get('start_date'), values.get('end_date')
                    if start_date and end_date:
                        if start_date == end_date:
                            condition = f"DATE({column_name}) = STR_TO_DATE('{start_date}', '%Y/%m/%d')"
                        else:
                            condition = f"{column_name} BETWEEN STR_TO_DATE('{start_date}', '%Y/%m/%d') AND STR_TO_DATE('{end_date}', '%Y/%m/%d')"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"日付条件: {condition}")
                elif input_fields_types[db_item] == 'FA' and values.strip():
                    condition = f"{column_name} LIKE '%{values}%'"
                    additional_conditions.append(condition)
                    LOGGER.debug(f"FA条件: {condition}")
                elif input_fields_types[db_item] == 'JSON' and isinstance(values, dict):
                    for path, value in values.items():
                        condition = f"JSON_CONTAINS({column_name}, '\"{value}\"', '$.{path}')"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"JSON条件 (dict): {condition}")
                elif input_fields_types[db_item] == 'JSON' and isinstance(values, str):
                    condition = f"JSON_CONTAINS({column_name}, '\"{values}\"', '$')"
                    additional_conditions.append(condition)
                    LOGGER.debug(f"JSON条件 (str): {condition}")
                elif values:
                    if isinstance(values, list):
                        placeholders = ', '.join([f"'{value}'" for value in values])
                        condition = f"{column_name} IN ({placeholders})"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"IN条件: {condition}")
                    else:
                        condition = f"{column_name} = '{values}'"
                        additional_conditions.append(condition)
                        LOGGER.debug(f"等価条件: {condition}")

        # カラム名からテーブルエイリアスを特定
        table_alias = find_table_alias(sql_query)

        # 削除除外条件の追加
        if not skip_deletion_exclusion:
            deletion_exclusion = str(deletion_exclusion).upper()
            if deletion_exclusion == 'TRUE':
                additional_conditions.append(table_alias + ".deleted_at IS NULL")
                LOGGER.debug("削除除外条件が追加されました")

        # SQLクエリの前処理
        sql_query = preprocess_sql_query(sql_query)

        # GROUP BY句の検出と除去
        sql_query, group_by_clause = detect_and_remove_group_by(sql_query)

        # WHERE句の存在チェックと追加
        sql_query = check_and_prepare_where_clause(sql_query, additional_conditions)

        # GROUP BY句を再追加
        if group_by_clause:
            sql_query += "\n" + group_by_clause
        else:
            sql_query += ";"
        # サブクエリを元に戻す
        sql_query = restore_subqueries(sql_query, subqueries)
        return sql_query
    except Exception as e:
        LOGGER.error(f"add_conditions_to_sql関数内でエラーが発生しました: {e}")
        raise

def set_period_condition(period_condition, period_criteria, sql_query, category):
    try:
        # 時間関連のカラム名を設定
        column_name = 'created_at' if period_criteria == '登録日時' else 'updated_at'

        # SQLクエリの前処理
        sql_query = preprocess_sql_query(sql_query)

        if category != 'マスタ':
            # GROUP BY句の検出と除去
            sql_query, group_by_clause = detect_and_remove_group_by(sql_query)

            # カラム名からテーブルエイリアスを特定
            table_alias = find_table_alias(sql_query)

            # 期間条件に基づくWHERE句の条件を生成
            condition = generate_period_condition(period_condition, column_name, table_alias)

            if condition:
                additional_conditions = [condition]
                # WHERE句の存在チェックと追加
                sql_query = check_and_prepare_where_clause(sql_query, additional_conditions)

            # GROUP BY句を再追加
            if group_by_clause:
                sql_query += "\n" + group_by_clause
            else:
                sql_query += ";"
        else:
            sql_query += ";"

        return sql_query
    except Exception as e:
        print(f"An error occurred: {e}")
        raise

#テーブルエイリアスを特定する関数
def find_table_alias(sql_query):
    # '-- FROM clause' コメント以降の部分を抽出
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        return None  # コメントが見つからない場合はNoneを返す
    
    # FROM句以降のテキストを取得
    sub_query = sql_query[from_clause_index + len('-- FROM clause'):]

    # 実際のFROM句を探す。ASキーワードの存在にも対応。
    from_match = re.search(r'\bFROM\b\s+(\w+)\s+(AS\s+)?(\w+)', sub_query, re.IGNORECASE)
    if from_match:
        # テーブル名
        table_name = from_match.group(1)
        # エイリアスが指定されている場合はそれを使用し、なければテーブル名をそのままエイリアスとして使用
        alias = from_match.group(3) if from_match.group(3) else table_name
        return alias
    return None

#SQLクエリの前処理を行う。セミコロンの削除とトリミングを含む。
def preprocess_sql_query(sql_query):
    if sql_query.strip().endswith(";"):
        sql_query = sql_query.strip()[:-1]
    return sql_query.strip()

#SQLクエリからGROUP BY句を検出し、取り除いてそれを返す。
def detect_and_remove_group_by(sql_query):
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        return sql_query, ""
    
    sub_query = sql_query[from_clause_index:]
    upper_sub_query = sub_query.upper()
    group_by_index = upper_sub_query.find("GROUP BY")
    if group_by_index != -1:
        group_by_clause = sub_query[group_by_index:]
        sub_query = sub_query[:group_by_index].strip()
        sql_query = sql_query[:from_clause_index] + sub_query
        return sql_query, group_by_clause
    return sql_query, ""

#サブクエリを検出
def detect_and_replace_subqueries(sql_query):
    subquery_pattern = r'-- subquery start(.*?)-- subquery end'
    subqueries = re.findall(subquery_pattern, sql_query, re.DOTALL)
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sql_query = sql_query.replace(f'-- subquery start{subquery}-- subquery end', placeholder)
    return sql_query, subqueries

#元に戻す
def restore_subqueries(sql_query, subqueries):
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sql_query = sql_query.replace(placeholder, f'-- subquery start{subquery}-- subquery end')
    return sql_query

#SQLクエリにWHERE句があるか確認し、適切な形で返す。
def check_and_prepare_where_clause(sql_query, additional_conditions):
    """WHERE句の存在を確認し、必要に応じてANDを追加する。追加条件が存在しない場合は何も追加しない。 """
    if not additional_conditions:
        return sql_query  # 条件がなければそのまま返す

    # "-- FROM clause"以降の部分を取得
    from_clause_index = sql_query.find('-- FROM clause')
    if from_clause_index == -1:
        raise ValueError("-- FROM clause not found in the SQL query")

    # FROM句の後の部分を取得
    sub_query = sql_query[from_clause_index + len('-- FROM clause\n'):].strip()

    # サブクエリを一時的に置換
    subquery_pattern = r'-- subquery start(.*?)-- subquery end'
    subqueries = re.findall(subquery_pattern, sub_query, re.DOTALL)
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        sub_query = sub_query.replace(f'-- subquery start{subquery}-- subquery end', placeholder)

    # WHERE句があるかどうかをチェック
    where_index = -1
    for match in re.finditer(r'\bWHERE\b', sub_query, re.IGNORECASE):
        if not re.search(r'__SUBQUERY_PLACEHOLDER_\d+__', sub_query[:match.start()]):
            where_index = match.start()
            break

    if where_index == -1:
        # WHERE句がない場合、WHERE句を追加
        modified_sub_query = sub_query + "\nWHERE " + " AND ".join(additional_conditions)
    else:
        # WHERE句がある場合、ANDで条件を追加
        where_clause = sub_query[:where_index + len("WHERE")]
        remaining_query = sub_query[where_index + len("WHERE"):].strip()
        modified_sub_query = where_clause + " " + remaining_query + " AND " + " AND ".join(additional_conditions)

    # サブクエリを元に戻す
    for i, subquery in enumerate(subqueries):
        placeholder = f'__SUBQUERY_PLACEHOLDER_{i}__'
        modified_sub_query = modified_sub_query.replace(placeholder, f'-- subquery start{subquery}-- subquery end')

    # 元のクエリのFROM句までの部分と、修正後のサブクエリを結合
    final_query = sql_query[:from_clause_index + len('-- FROM clause\n')] + "\n" + modified_sub_query
    return final_query

def generate_period_condition(period_condition, column_name, table_alias):
    """期間条件に基づくWHERE句の条件を生成する。"""
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    date_column = f"{table_alias}.{column_name}" if table_alias else column_name
    if period_condition == '当日':
        return f" DATE({date_column}) = '{today}'"
    elif period_condition == '前日':
        return f" DATE({date_column}) = '{yesterday}'"
    elif period_condition == '前日まで累積':
        return f" DATE({date_column}) <= '{yesterday}'"
    elif period_condition == '当日まで累積':
        return f" DATE({date_column}) <= '{today}'"
    elif '～前日まで累積' in period_condition:
        start_date_str = period_condition.split('～')[0].strip()
        start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
        return f" DATE({date_column}) BETWEEN '{start_date}' AND '{yesterday}'"
    elif '～当日まで累積' in period_condition:
        start_date_str = period_condition.split('～')[0].strip()
        start_date = datetime.strptime(start_date_str, '%Y年%m月%d日').date()
        return f" DATE({date_column}) BETWEEN '{start_date}' AND '{today}'"
    elif '～' in period_condition:
        parts = period_condition.split('～')
        if len(parts) == 3 and parts[2].strip() == 'までの期間':
            start_date_str, end_date_str = parts[0:2]
            start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
            end_date = datetime.strptime(end_date_str.strip(), '%Y年%m月%d日').date()
            return f" DATE({date_column}) BETWEEN '{start_date}' AND '{end_date}'"
        elif 'まで累積' not in period_condition:
            start_date_str, end_date_str = parts
            start_date = datetime.strptime(start_date_str.strip(), '%Y年%m月%d日').date()
            end_date = datetime.strptime(end_date_str.strip(), '%Y年%m月%d日').date()
            return f" DATE({date_column}) BETWEEN '{start_date}' AND '{end_date}'"
    elif '日前時点を1日分' in period_condition:
        days_ago = int(period_condition.split('日前')[0])
        target_date = today - timedelta(days=days_ago)
        return f" DATE({date_column}) = '{target_date}'"
    elif '年' in period_condition and '月' in period_condition and '日' in period_condition:
        # 特定の日付（YYYY年MM月DD日）の場合
        specific_date = datetime.strptime(period_condition.strip(), '%Y年%m月%d日').date()
        return f" DATE({date_column}) = '{specific_date}'"
    else:
        return ""

def get_column_letter(column_index):
    # 1から始まるインデックスに対応するため、1を引く
    column_index -= 1
    letter = ''
    while column_index >= 0:
        letter = chr(column_index % 26 + 65) + letter
        column_index = column_index // 26 - 1
    return letter

#スプシ貼り付け
@retry_on_exception
def export_to_spreadsheet(conn, sql_query, save_path_id, sheet_name, json_keyfile_path, paste_format, main_sheet_name, csv_file_name_column, main_table_name, category, chunk_size=10000, delay=0):

    print("SQLクエリの実行を開始します。（スプシの貼り付け）")
    cursor = conn.cursor()
    try:
        cursor.execute(sql_query)
        fetchall = cursor.fetchall()
        record_count = len(fetchall)

        # ヘッダ行の取得
        headers = [i[0] for i in cursor.description]

        # Decimalおよびdatetimeオブジェクトを文字列に変換しない
        converted_data = []
        for row in fetchall:
            converted_row = [str(cell) if isinstance(cell, Decimal) else cell if cell is not None else '' for cell in row]
            converted_data.append(converted_row)
        
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "成功", None, save_path_id)  # ログシートに成功を書き込む

        SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = authenticate_google_api(json_keyfile_path, SCOPES)
        gc = gspread.authorize(credentials)

        spreadsheet = gc.open_by_key(save_path_id)

        if sheet_name == main_sheet_name:
            # 実行シートの場合は処理をスキップ
            print(f"Skipping export to main sheet: {sheet_name}")
            return

        if not sheet_name:
            # "CSVファイル名/SSシート名" がブランクの場合、"CSVファイル呼称" 列の値を使用
            sheet_name = csv_file_name_column
            print("CSVファイル名/SSシート名がブランクの場合、")
            print(csv_file_name_column)

        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            # 指定されたシートが存在しない場合は新しいシートを追加
            print("シート作成")
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=26)
            # 新しいシートの場合、ヘッダ行を追加し、ブランク行を追加
            worksheet.update('A1', [headers])
            worksheet.append_row([''] * len(headers))  # ブランク行の追加

        # 既存シートの1行目をチェックし、ヘッダ行が存在しない場合は追加
        if not worksheet.row_values(1):
            worksheet.update('A1', [headers])
            worksheet.append_row([''] * len(headers))  # ブランク行の追加
        else:
            print("ヘッダ行は既に存在します。")

        column_count = len(cursor.description)
        last_column_letter = get_column_letter(column_count)
        print(f"Column count: {column_count}, Last column letter: {last_column_letter}")

        if paste_format == '最終行積立て':
            last_row = len(worksheet.col_values(1)) + 1
            if last_row > worksheet.row_count:
                additional_rows = last_row - worksheet.row_count
                worksheet.add_rows(additional_rows)
            for i in range(0, len(converted_data), chunk_size):
                chunk = converted_data[i:i + chunk_size]
                worksheet.update(f'A{last_row}', chunk)
                last_row += len(chunk)
                if delay:
                    print(f"{delay}秒待機します。")
                    time.sleep(delay)
        elif paste_format == '全張替え':
            # 既存のヘッダ行を保持し、データ部分のみクリア
            clear_range = f'A2:{last_column_letter}{worksheet.row_count}'
            print(f"Clearing range: {clear_range}")
            worksheet.batch_clear([clear_range])

            data_row_count = len(converted_data)
            rows_to_add = data_row_count - (worksheet.row_count - 1)
            # 1万行ずつ追加
            while rows_to_add > 0:
                worksheet.add_rows(min(10000, rows_to_add))
                rows_to_add -= 10000

            row_start = 2  # データの開始行は2行目から
            for i in range(0, len(converted_data), chunk_size):
                chunk = converted_data[i:i + chunk_size]
                worksheet.update(f'A{row_start}', chunk)
                row_start += len(chunk)
                if delay:
                    print(f"{delay}秒待機します。")
                    time.sleep(delay)

        print(f"Data has been transferred to {sheet_name} sheet in {save_path_id} with {paste_format} method.")
    except Exception as e:
        print(f"Error during query execution or spreadsheet writing: {e}")
        write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, "失敗", str(e), save_path_id)  # ログシートに失敗を書き込む
        raise
    finally:
        cursor.close()

#テスト実行
@retry_on_exception
def setup_test_environment(test_execution, output_to_spreadsheet, save_path_id, csv_file_name, spreadsheet_id, json_keyfile_path):
    if str(test_execution).lower() == 'true':
        if output_to_spreadsheet == 'CSV':
            test_folder = os.path.join(save_path_id, 'test')
            if not os.path.exists(test_folder):
                os.makedirs(test_folder)
            save_path_id = test_folder
        elif output_to_spreadsheet == 'スプシ':
            SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            credentials = authenticate_google_api(json_keyfile_path, SCOPES)
            gc = gspread.authorize(credentials)
            
            spreadsheet_id = save_path_id
            
            spreadsheet = gc.open_by_key(save_path_id)
            
            test_sheet_name = f"{csv_file_name}_test"
            try:
                worksheet = spreadsheet.worksheet(test_sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                try:
                    source_sheet = spreadsheet.worksheet(csv_file_name)
                    worksheet = spreadsheet.add_worksheet(title=test_sheet_name, rows=source_sheet.row_count, cols=source_sheet.col_count)
                    header_row = source_sheet.row_values(1)
                    worksheet.insert_row(header_row, index=1)
                except gspread.exceptions.WorksheetNotFound:
                    raise ValueError(f"The source sheet '{csv_file_name}' does not exist in the spreadsheet.")
            csv_file_name = test_sheet_name

    return save_path_id, csv_file_name

#ログの書き出し
def write_to_log_sheet(csv_file_name_column, sheet_name, main_table_name, category, record_count, json_keyfile_path, result, error_log=None, save_path_id=None):
    log_spreadsheet_id = '1iqDqeGXAovNQfnuuOi2xLzJIrmXOE1FKOgrSLgG0SOw'
    log_sheet_name = 'log'

    SCOPES = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = authenticate_google_api(json_keyfile_path, SCOPES)
    client = gspread.authorize(credentials)

    try:
        spreadsheet = client.open_by_key(log_spreadsheet_id)
        worksheet = spreadsheet.worksheet(log_sheet_name)

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [csv_file_name_column, sheet_name, main_table_name, category, save_path_id, record_count, result, error_log, timestamp, '']
        worksheet.append_row(row_data)

        print(f"ログシートに書き込みました: {row_data}")
    except gspread.exceptions.APIError as e:
        print(f"ログシートへの書き込み中にエラーが発生しました: {e}")

# DB項目のデータ型を強制指定
def get_data_types(worksheet):
    headers = worksheet.row_values(1)
    data_types = {}

    db_item_col = None
    data_type_col = None

    for i, header in enumerate(headers):
        if header == 'DB項目':
            db_item_col = i
        elif header == 'DATA_TYPE':
            data_type_col = i

    if db_item_col is not None and data_type_col is not None:
        last_row = len(worksheet.col_values(db_item_col + 1))

        db_item_range = worksheet.range(2, db_item_col + 1, last_row, db_item_col + 1)
        data_type_range = worksheet.range(2, data_type_col + 1, last_row, data_type_col + 1)

        db_items = [cell.value for cell in db_item_range]
        data_types_list = [cell.value for cell in data_type_range]

        data_types = dict(zip(db_items, data_types_list))
    
    return data_types

# データ型を適用する関数（フォーマット済み）
def apply_data_types_to_df(df, data_types, LOGGER, encoding='utf-8'):
    converted_columns = []  # 型変換を行った列名を格納するリスト
    for column, data_type in data_types.items():
        if column in df.columns:
            try:
                if data_type == 'txt':
                    df[column] = df[column].apply(lambda x: x.encode(encoding).decode(encoding) if isinstance(x, str) else x)
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'int':
                    df[column] = pd.to_numeric(df[column], errors='raise').astype('Int64')
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'float':
                    df[column] = pd.to_numeric(df[column], errors='raise').astype(float)
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'date':
                    df[column] = df[column].astype(str)  # date型は文字列に変換済み
                    converted_columns.append(column)  # 型変換を行った列名を追加
                elif data_type == 'datetime':
                    df[column] = df[column].astype(str)  # datetime型は文字列に変換済み
                    converted_columns.append(column)  # 型変換を行った列名を追加
            except ValueError as e:
                LOGGER.error(f"Error converting column '{column}' to type '{data_type}': {e}")
                raise ValueError(f"Error converting column '{column}' to type '{data_type}': {e}")
            except Exception as e:
                LOGGER.error(f"Unexpected error processing column '{column}': {e}")
                raise Exception(f"Unexpected error processing column '{column}': {e}")
    
    if converted_columns:
        LOGGER.info(f"以下の列の型変換を行いました: {', '.join(converted_columns)}")
    else:
        LOGGER.info("型変換は行われませんでした。")
    
    return df

#スプシに基づき編集するSQL
def execute_sql_query_with_conditions(sql_file_name, config, period_condition, period_criteria, deletion_exclusion, category, main_table_name):
    sql_query = load_sql_from_file(sql_file_name, config['google_folder_id'], config['json_keyfile_path'])
    if sql_query:
        try:
            input_values, input_fields_types = {}, {}
            sql_query_with_period_condition = set_period_condition(period_condition, period_criteria, sql_query, category)
            if category != 'マスタ':
                sql_query_with_conditions = add_conditions_to_sql(sql_query_with_period_condition, input_values, input_fields_types, deletion_exclusion)
            else:
                sql_query_with_conditions = sql_query_with_period_condition
            return sql_query_with_conditions
        except Exception as e:
            LOGGER.error(f"SQLクエリの処理中にエラーが発生しました: {e}")
            return None
    else:
        LOGGER.warning(f"{sql_file_name} の読み込みに失敗しました。代わりに 'SELECT *' を実行します。")
        return f"SELECT * -- FROM clause\nFROM {main_table_name}"